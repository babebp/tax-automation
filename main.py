from io import BytesIO
from copy import copy
from urllib.parse import quote
from googleapiclient.http import MediaIoBaseDownload
import google.generativeai as genai 
import logging
from typing import List, Optional, Dict
from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import sqlite3
from contextlib import contextmanager
from fastapi.responses import StreamingResponse
import openpyxl
from dotenv import load_dotenv
import os
import re
import pandas as pd
import requests
from datetime import datetime


load_dotenv() # Load environment variables from .env file

import google_drive as gd
# ... (rest of the file)

# Configure logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()])

try:
    from openpyxl.formula.translate import Translator
    CAN_TRANSLATE = True
except ImportError:
    CAN_TRANSLATE = False
    logging.warning("Could not import openpyxl.formula.translate.Translator. Formulas will be copied without adjusting references.")

# Configure Gemini API
# IMPORTANT: The API key should be set in your environment or a secure config
# For this project, we assume it's in .streamlit/secrets.toml and loaded by the environment
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)
else:
    logging.warning("GEMINI_API_KEY environment variable not set. PDF processing will fail.")

# Configure LINE API
LINE_API_URL = "https://api.line.me/v2/bot/message/push"

DB_PATH = "app.db"

# ... (Database helpers and Pydantic schemas remain the same) ...
@contextmanager
def get_conn():
    """Context manager to handle database connection."""
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.row_factory = sqlite3.Row
        yield conn
        conn.commit()
    finally:
        conn.close()

def init_db():
    """Initializes the database tables if they don't exist."""
    with get_conn() as conn:
        cur = conn.cursor()
        # companies table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            google_drive_folder_id TEXT,
            google_drive_folder_name TEXT
        )
        """)
        # banks table (many-to-one relationship with companies)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS company_banks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            bank_name TEXT NOT NULL,
            tb_code TEXT NOT NULL,
            UNIQUE(company_id, bank_name),
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
        )
        """)
        # forms table (many-to-one relationship with companies)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS company_forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_id INTEGER NOT NULL,
            form_type TEXT NOT NULL,
            tb_code TEXT NOT NULL,
            UNIQUE(company_id, form_type),
            FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
        )
        """)
        # Drop the old line_recipients table if it exists for a clean slate
        cur.execute("DROP TABLE IF EXISTS line_recipients")
        
        # Drop the old user table to redefine it
        cur.execute("DROP TABLE IF EXISTS line_users")
        
        # Create a new, global user table without channel_id
        cur.execute("""
        CREATE TABLE IF NOT EXISTS line_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            uid TEXT UNIQUE NOT NULL,
            display_name TEXT NOT NULL
        )
        """)
        # line_channels table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS line_channels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL,
            token TEXT NOT NULL,
            channel_id_line TEXT
        )
        """)
        # line_groups table
        cur.execute("""
        CREATE TABLE IF NOT EXISTS line_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_id TEXT NOT NULL,
            group_name TEXT,
            channel_id INTEGER NOT NULL,
            UNIQUE(group_id, channel_id),
            FOREIGN KEY(channel_id) REFERENCES line_channels(id) ON DELETE CASCADE
        )
        """)

# ---------- Pydantic schemas ----------
class CompanyCreate(BaseModel):
    name: str = Field(..., min_length=1)

class LineEventSource(BaseModel):
    userId: Optional[str] = None
    groupId: Optional[str] = None
    roomId: Optional[str] = None

class LineEvent(BaseModel):
    type: str
    source: LineEventSource
    destination: Optional[str] = None

class LineWebhook(BaseModel):
    events: List[LineEvent]
    destination: Optional[str] = None # For top-level destination

class Company(BaseModel):
    id: int
    name: str
    google_drive_folder_id: Optional[str] = None
    google_drive_folder_name: Optional[str] = None

class BankCreate(BaseModel):
    company_id: int
    bank_name: str = Field(..., min_length=1)
    tb_code: str = Field(..., min_length=1)

class Bank(BaseModel):
    id: int
    company_id: int
    bank_name: str
    tb_code: str

class FormsUpsert(BaseModel):
    data: Dict[str, str]

class WorkflowStart(BaseModel):
    company_id: int
    month: str
    year: int

class ReconcileStart(BaseModel):
    company_id: int
    month: str
    year: int
    parts: List[str]

class LineChannelCreate(BaseModel):
    name: str = Field(..., min_length=1)
    token: str = Field(..., min_length=1)

class LineChannel(BaseModel):
    id: int
    name: str
    token: str

class LineMessageSend(BaseModel):
    channel_id: int
    message: str = Field(..., min_length=1)
    recipient_uids: List[str] = []
    recipient_gids: List[str] = []

class LineUser(BaseModel):
    id: int
    uid: str
    display_name: str

class LineGroup(BaseModel):
    id: int
    group_id: str
    group_name: str
    channel_id: int

# ---------- New Helper Functions for PDF and LLM ----------
def get_amount_from_gemini(file_content: bytes, prompt: str) -> str:
    """Sends a PDF file to Gemini LLM for OCR and returns the extracted amount."""
    if not GEMINI_API_KEY:
        return "GEMINI_API_KEY not set"
    if not file_content:
        return "No file content"

    try:
        model = genai.GenerativeModel('gemini-2.5-flash')
        response = model.generate_content([prompt, {"mime_type": "application/pdf", "data": file_content}])
        return response.text.strip()
    except Exception as e:
        logging.error(f"Gemini API call failed: {e}")
        return f"Error: {e}"

def clean_and_convert_to_float(text: str) -> Optional[float]:
    """Cleans a string to remove non-numeric characters and converts it to a float."""
    if not isinstance(text, str):
        return None
    try:
        # Remove commas, currency symbols, and any non-numeric characters except for the decimal point and minus sign.
        cleaned_text = re.sub(r'[^\d.-]', '', text)
        if cleaned_text:
            return float(cleaned_text)
    except (ValueError, TypeError):
        return None
    return None

# ---------- Constants ----------
FIXED_FORMS = ["PND1", "PND3", "PND53", "PP30", "SSO", "Revenue", "Revenue2", "Credit Note"]

PROMPTS = {
    "BANK": "หายอดคงเหลืสุดท้ายให้หน่อย ตอบมาแค่ตัวเลขเท่านั้น ห้ามมีตัวหนังสืออื่นๆเด็ดขาด",
    "PND1": "ภายในหัวข้อ \"สำหรับใบเสร็จรับเงิน\" ให้ตัวเลขของ \"จำนวนเงิน\" ออกมา ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามมีตัวหนังสือเด็ดขาด",
    "PND3": "ภายในหัวข้อ \"สำหรับใบเสร็จรับเงิน\" ให้ตัวเลขของ \"จำนวนเงิน\" ออกมา ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามมีตัวหนังสือเด็ดขาด",
    "PND53": "ภายในหัวข้อ \"สำหรับใบเสร็จรับเงิน\" ให้ตัวเลขของ \"จำนวนเงิน\" ออกมา ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามมีตัวหนังสือเด็ดขาด",
    "PP30": """ภายในหัวข้อ \"ภาษีสุทธิ\" ให้ดึงตัวเลขของช่องที่ 11 หรือช่องที่ 12 มาโดยหากดึงจากช่อง 11 ให้ดึงมาตรงๆ แต่ถ้าเป็นช่องที่ 12 ให้ดึงแล้วเปลี่ยนเป็นค่าลบ ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามที่ตัวหนังสือเด็ดขาด\n\nหมายเหตุ: ในเอกสารจะมีเลขแค่ช่องใดช่องหนึ่งเท่านั้น (ไม่ 11 ก็ 12)""",
    "SSO": "ภายในหัวข้อ \"จำนวนเงินที่ำระ\" ให้ตัวเลออกมา ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามมีตัวหนังสือเด็ดขาด"
}

THAI_MONTHS = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]

# ... (FastAPI app setup and other endpoints remain the same) ...
app = FastAPI(title="Company Settings API", version="1.0.0")


origins = [
    "https://app.byteduck.io",
    "http://localhost:8501"
]
# Enable CORS for frontend development
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def on_startup():
    """Initialize the database when the application starts."""
    init_db()

# ---------- Company endpoints ----------
@app.get("/companies", response_model=List[Company])
def list_companies():
    """Lists all companies."""
    with get_conn() as conn:
        rows = conn.execute("SELECT id, name, google_drive_folder_id, google_drive_folder_name FROM companies ORDER BY name").fetchall()
        return [Company(id=r["id"], name=r["name"], google_drive_folder_id=r["google_drive_folder_id"], google_drive_folder_name=r["google_drive_folder_name"]) for r in rows]

@app.post("/companies", response_model=Company)
def create_company(payload: CompanyCreate):
    """Creates a new company."""
    with get_conn() as conn:
        try:
            cur = conn.execute("INSERT INTO companies(name) VALUES (?)", (payload.name.strip(),))
            cid = cur.lastrowid
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Company name already exists.")
        row = conn.execute("SELECT id, name FROM companies WHERE id = ?", (cid,)).fetchone()
        return Company(id=row["id"], name=row["name"])

@app.delete("/companies/{company_id}")
def delete_company(company_id: int):
    """Deletes a company and all its associated data."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Company not found.")
        conn.execute("DELETE FROM companies WHERE id = ?", (company_id,))
        return {"ok": True}

@app.put("/companies/{company_id}", response_model=Company)
def update_company(company_id: int, payload: CompanyCreate):
    """Updates a company's name."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Company not found.")
        try:
            conn.execute("UPDATE companies SET name = ? WHERE id = ?", (payload.name.strip(), company_id))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=400, detail="Company name already exists.")
        row = conn.execute("SELECT id, name, google_drive_folder_id, google_drive_folder_name FROM companies WHERE id = ?", (company_id,)).fetchone()
        return Company(id=row["id"], name=row["name"], google_drive_folder_id=row["google_drive_folder_id"], google_drive_folder_name=row["google_drive_folder_name"])

class CompanyDriveFolderUpdate(BaseModel):
    google_drive_folder_id: str
    google_drive_folder_name: str

@app.put("/companies/{company_id}/google-drive-folder", response_model=Company)
def update_company_drive_folder(company_id: int, payload: CompanyDriveFolderUpdate):
    """Updates the Google Drive folder for a company."""
    with get_conn() as conn:
        c = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not c:
            raise HTTPException(status_code=404, detail="Company not found.")
        conn.execute(
            "UPDATE companies SET google_drive_folder_id = ?, google_drive_folder_name = ? WHERE id = ?",
            (payload.google_drive_folder_id, payload.google_drive_folder_name, company_id)
        )
        row = conn.execute("SELECT id, name, google_drive_folder_id, google_drive_folder_name FROM companies WHERE id = ?", (company_id,)).fetchone()
        return Company(id=row["id"], name=row["name"], google_drive_folder_id=row["google_drive_folder_id"], google_drive_folder_name=row["google_drive_folder_name"])

# ---------- Bank endpoints ----------
@app.get("/companies/{company_id}/banks", response_model=List[Bank])
def list_banks(company_id: int):
    """Lists all banks for a given company."""
    with get_conn() as conn:
        c = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not c:
            raise HTTPException(status_code=404, detail="Company not found.")
        rows = conn.execute("SELECT id, company_id, bank_name, tb_code FROM company_banks WHERE company_id = ? ORDER BY bank_name", (company_id,)).fetchall()
        return [Bank(id=r["id"], company_id=r["company_id"], bank_name=r["bank_name"], tb_code=r["tb_code"]) for r in rows]

@app.post("/banks", response_model=Bank)
def add_bank(payload: BankCreate):
    """Adds a new bank to a company and assigns it the next available index."""
    with get_conn() as conn:
        c = conn.execute("SELECT id FROM companies WHERE id = ?", (payload.company_id,)).fetchone()
        if not c:
            raise HTTPException(status_code=404, detail="Company not found.")
        
        cur = conn.execute(
            "INSERT INTO company_banks(company_id, bank_name, tb_code) VALUES (?, ?, ?)",
            (payload.company_id, payload.bank_name.strip(), payload.tb_code.strip())
        )
        bid = cur.lastrowid
        row = conn.execute("SELECT id, company_id, bank_name, tb_code FROM company_banks WHERE id = ?", (bid,)).fetchone()
        return Bank(id=row["id"], company_id=row["company_id"], bank_name=row["bank_name"], tb_code=row["tb_code"])

@app.delete("/banks/{bank_id}")
def delete_bank(bank_id: int):
    """Deletes a bank."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM company_banks WHERE id = ?", (bank_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Bank record not found.")
        conn.execute("DELETE FROM company_banks WHERE id = ?", (bank_id,))
        return {"ok": True}

# ---------- Forms endpoints ----------
@app.get("/companies/{company_id}/forms")
def get_company_forms(company_id: int):
    """Gets all form configurations for a company."""
    with get_conn() as conn:
        c = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not c:
            raise HTTPException(status_code=404, detail="Company not found.")
        rows = conn.execute("SELECT form_type, tb_code FROM company_forms WHERE company_id = ?", (company_id,)).fetchall()
        existing = {r["form_type"]: r["tb_code"] for r in rows}
        result = {ft: existing.get(ft, "") for ft in FIXED_FORMS}
        return {"company_id": company_id, "forms": result, "fixed": FIXED_FORMS}

@app.put("/companies/{company_id}/forms")
def upsert_company_forms(company_id: int, payload: FormsUpsert):
    """Updates or creates form configurations for a company."""
    for k in payload.data.keys():
        if k not in FIXED_FORMS:
            raise HTTPException(status_code=400, detail=f"Invalid form type: {k}")
    with get_conn() as conn:
        c = conn.execute("SELECT id FROM companies WHERE id = ?", (company_id,)).fetchone()
        if not c:
            raise HTTPException(status_code=404, detail="Company not found.")
        for form_type, tb_code in payload.data.items():
            tb_code = tb_code.strip()
            conn.execute("INSERT INTO company_forms(company_id, form_type, tb_code) VALUES (?, ?, ?) ON CONFLICT(company_id, form_type) DO UPDATE SET tb_code=excluded.tb_code", (company_id, form_type, tb_code))
        return {"ok": True}

# ---------- LINE Notify endpoints ----------
@app.post("/line/webhook")
def line_webhook(payload: LineWebhook):
    """Handles incoming LINE messages to capture user information."""
    with get_conn() as conn:
        all_channels = conn.execute("SELECT id, name, token FROM line_channels").fetchall()
        if not all_channels:
            logging.error("Webhook called, but no channels are configured.")
            return {"ok": False, "detail": "No channels configured"}

    for event in payload.events:
        logging.info(f"Received LINE event: type={event.type}, userId={event.source.userId}, groupId={event.source.groupId}")

        # --- New Logic: Iterate through channels and test token viability ---
        target_channel = None
        user_id = event.source.userId
        chat_id = event.source.groupId if event.source.groupId else event.source.roomId

        # Try to find the correct channel by checking which one can access the user/group info
        for ch in all_channels:
            access_token = ch["token"]
            headers = {"Authorization": f"Bearer {access_token}"}
            
            # If it's a user event, try fetching the user's profile
            if user_id and not chat_id:
                profile_url = f"https://api.line.me/v2/bot/profile/{user_id}"
                try:
                    res = requests.get(profile_url, headers=headers, timeout=2)
                    if res.status_code == 200:
                        target_channel = ch
                        logging.info(f"Successfully validated token for channel '{ch['name']}' (ID: {ch['id']}) with user {user_id}.")
                        break # Found the correct channel
                except requests.RequestException:
                    continue # Try next token
            
            # If it's a group event, try fetching group summary
            elif chat_id:
                summary_url = f"https://api.line.me/v2/bot/group/{chat_id}/summary"
                try:
                    res = requests.get(summary_url, headers=headers, timeout=2)
                    if res.status_code == 200:
                        target_channel = ch
                        logging.info(f"Successfully validated token for channel '{ch['name']}' (ID: {ch['id']}) with group {chat_id}.")
                        break # Found the correct channel
                except requests.RequestException:
                    continue # Try next token

        if not target_channel:
            logging.warning(f"Could not find a channel with a valid token to handle event for user '{user_id}' or group '{chat_id}'.")
            continue # Skip to the next event

        # --- Proceed with the identified channel ---
        registered_channel_id = target_channel["id"]
        access_token = target_channel["token"]
        headers = {"Authorization": f"Bearer {access_token}"}

        # Handle user messages (direct 1-on-1 chat)
        if event.type == "message" and user_id:
            profile_url = f"https://api.line.me/v2/bot/profile/{user_id}"
            try:
                res = requests.get(profile_url, headers=headers, timeout=3)
                res.raise_for_status()
                profile = res.json()
                display_name = profile.get("displayName", "Unknown")
                
                with get_conn() as conn:
                    conn.execute(
                        "INSERT INTO line_users (uid, display_name) VALUES (?, ?) ON CONFLICT(uid) DO UPDATE SET display_name=excluded.display_name",
                        (user_id, display_name)
                    )
                logging.info(f"Successfully registered/updated global user '{display_name}' ({user_id}).")
            except requests.RequestException as e:
                logging.error(f"Could not fetch LINE profile for UID {user_id} on channel {registered_channel_id}: {e}")
            
        # Handle bot joining a group or room
        elif event.type == "join" and chat_id:
            summary_url = f"https://api.line.me/v2/bot/group/{chat_id}/summary"
            try:
                res = requests.get(summary_url, headers=headers, timeout=3)
                res.raise_for_status()
                summary = res.json()
                chat_name = summary.get("groupName", f"Group ({chat_id})")
                with get_conn() as conn:
                    conn.execute(
                        "INSERT INTO line_groups (group_id, group_name, channel_id) VALUES (?, ?, ?) ON CONFLICT(group_id, channel_id) DO UPDATE SET group_name=excluded.group_name",
                        (chat_id, chat_name, registered_channel_id)
                    )
                logging.info(f"Bot from channel '{target_channel['name']}' joined group '{chat_name}' ({chat_id}).")
            except requests.RequestException as e:
                logging.error(f"Could not fetch LINE group summary for GID {chat_id} on channel {registered_channel_id}: {e}")
            
        # Handle bot leaving a group or room
        elif event.type == "leave" and chat_id:
            logging.info(f"Bot left chat: {chat_id} for channel {registered_channel_id}. Removing from DB.")
            with get_conn() as conn:
                conn.execute("DELETE FROM line_groups WHERE group_id = ? AND channel_id = ?", (chat_id, registered_channel_id))

    return {"ok": True}

@app.get("/line/users", response_model=List[LineUser])
def list_line_users():
    """Lists all unique LINE users who have interacted with any bot."""
    with get_conn() as conn:
        rows = conn.execute("SELECT id, uid, display_name FROM line_users ORDER BY display_name").fetchall()
        return [LineUser(id=r["id"], uid=r["uid"], display_name=r["display_name"]) for r in rows]


@app.get("/line/groups", response_model=List[LineGroup])
def list_line_groups(channel_id: Optional[int] = None):
    """Lists all groups the LINE bot is currently a member of."""
    with get_conn() as conn:
        if channel_id:
            rows = conn.execute("SELECT id, group_id, group_name, channel_id FROM line_groups WHERE channel_id = ? ORDER BY group_name", (channel_id,)).fetchall()
        else:
            rows = conn.execute("SELECT id, group_id, group_name, channel_id FROM line_groups ORDER BY group_name").fetchall()
        return [LineGroup(id=r["id"], group_id=r["group_id"], group_name=r["group_name"], channel_id=r["channel_id"]) for r in rows]


# ---------- LINE Channel endpoints ----------
@app.get("/line/channels", response_model=List[LineChannel])
def list_line_channels():
    """Lists all LINE channels."""
    with get_conn() as conn:
        rows = conn.execute("SELECT id, name, token FROM line_channels ORDER BY name").fetchall()
        return [LineChannel(id=r["id"], name=r["name"], token=r["token"]) for r in rows]

@app.post("/line/channels", response_model=LineChannel)
def add_line_channel(payload: LineChannelCreate):
    """Adds a new LINE channel."""
    # 1. Verify token and get bot info
    bot_info_url = "https://api.line.me/v2/bot/info"
    token = payload.token.strip()
    name = payload.name.strip()
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(bot_info_url, headers=headers, timeout=5)
        res.raise_for_status()
        bot_info = res.json()
        channel_id_line = bot_info.get("userId")
        if not channel_id_line:
            raise HTTPException(status_code=400, detail="Could not retrieve bot user ID from token. Is it a valid Messaging API token?")
    except requests.HTTPError as e:
        raise HTTPException(status_code=400, detail=f"Invalid token. LINE API returned status {e.response.status_code}.")
    except requests.RequestException as e:
        logging.error(f"Failed to verify LINE token: {e}")
        raise HTTPException(status_code=500, detail="Could not contact LINE API to verify token.")

    # 2. Add to database
    with get_conn() as conn:
        # Check if bot user ID already exists
        existing_bot = conn.execute("SELECT id FROM line_channels WHERE channel_id_line = ?", (channel_id_line,)).fetchone()
        if existing_bot:
            raise HTTPException(status_code=400, detail="A channel for this bot already exists.")
        
        try:
            cur = conn.execute(
                "INSERT INTO line_channels(name, token, channel_id_line) VALUES (?, ?, ?)", 
                (name, token, channel_id_line)
            )
            cid = cur.lastrowid
        except sqlite3.IntegrityError:
            # This will catch duplicate names
            raise HTTPException(status_code=400, detail="Channel name already exists.")
        
        row = conn.execute("SELECT id, name, token FROM line_channels WHERE id = ?", (cid,)).fetchone()
        return LineChannel(id=row["id"], name=row["name"], token=row["token"])

@app.delete("/line/channels/{channel_id}")
def delete_line_channel(channel_id: int):
    """Deletes a LINE channel."""
    with get_conn() as conn:
        row = conn.execute("SELECT id FROM line_channels WHERE id = ?", (channel_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Channel not found.")
        conn.execute("DELETE FROM line_channels WHERE id = ?", (channel_id,))
        return {"ok": True}


@app.post("/line/send_message")
def send_line_message(payload: LineMessageSend):
    """Sends a message to specified LINE recipients using a specific channel."""
    with get_conn() as conn:
        # Get the channel's access token
        channel_row = conn.execute("SELECT token FROM line_channels WHERE id = ?", (payload.channel_id,)).fetchone()
        if not channel_row:
            raise HTTPException(status_code=404, detail="LINE channel not found.")
        access_token = channel_row["token"]

    uids_to_send = payload.recipient_uids + payload.recipient_gids

    if not uids_to_send:
        raise HTTPException(status_code=400, detail="No recipients selected.")

    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}"
    }
    
    error_details = []
    success_count = 0

    for uid in uids_to_send:
        body = {
            "to": uid,
            "messages": [{"type": "text", "text": payload.message}]
        }
        try:
            response = requests.post(LINE_API_URL, headers=headers, json=body)
            response.raise_for_status()
            success_count += 1
        except requests.HTTPError as e:
            logging.error(f"Failed to send message to {uid}: {e.response.text}")
            error_details.append(f"Failed for UID {uid}: {e.response.status_code}")
    
    if error_details:
        raise HTTPException(
            status_code=500, 
            detail=f"Sent {success_count}/{len(uids_to_send)} messages. Errors: {', '.join(error_details)}"
        )

    return {"ok": True, "sent_count": success_count}

class GoogleDriveFolder(BaseModel):
    id: str
    name: str

@app.get("/google-drive/folders", response_model=List[GoogleDriveFolder])
def list_google_drive_folders(parent_folder_name: Optional[str] = None):
    """
    Lists folders. If parent_folder_name is provided, lists direct child folders.
    Otherwise, lists all accessible folders.
    """
    try:
        drive_service = gd.get_drive_service()
        
        if parent_folder_name and parent_folder_name.strip():
            # 1. Find parent folder ID
            # Ensure the name is properly escaped for the query
            safe_name = parent_folder_name.strip().replace("'", "\'\'")
            parent_query = f"name = '{safe_name}' and mimeType = 'application/vnd.google-apps.folder'"
            parent_folders = gd.find_files(drive_service, parent_query)
            
            if not parent_folders:
                logging.warning(f"Parent folder '{parent_folder_name}' not found.")
                return [] # Return empty list if parent not found
            
            parent_folder_id = parent_folders[0]['id']
            logging.info(f"Found parent folder '{parent_folder_name}' with ID: {parent_folder_id}")
            
            # 2. Find child folders
            query = f"'{parent_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
        else:
            # Default behavior: find all folders
            query = "mimeType = 'application/vnd.google-apps.folder'"

        folders = gd.find_files(drive_service, query)
        return [GoogleDriveFolder(id=f['id'], name=f['name']) for f in folders]
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Google Drive API error: {e}")


# ---------- Workflow endpoints ----------
@app.post("/workflow/start")
def start_workflow(payload: WorkflowStart):
    # ... (Existing setup code: logging, get company settings, auth with Drive) ...
    logging.info(f"Workflow started for company_id: {payload.company_id}, month: {payload.month}, year: {payload.year}")
    
    # Get company settings from the database
    logging.info("Fetching company settings from database.")
    with get_conn() as conn:
        c = conn.execute("SELECT name, google_drive_folder_id FROM companies WHERE id = ?", (payload.company_id,)).fetchone()
        if not c:
            logging.error(f"Company with id {payload.company_id} not found.")
            raise HTTPException(status_code=404, detail="Company not found.")
        company_name = c["name"]
        company_folder_id = c["google_drive_folder_id"]
        if not company_folder_id:
            raise HTTPException(status_code=400, detail="Company does not have a Google Drive folder configured.")
        banks_cursor = conn.execute("SELECT bank_name, tb_code FROM company_banks WHERE company_id = ? ORDER BY bank_name", (payload.company_id,))
        banks = banks_cursor.fetchall()
        forms_cursor = conn.execute("SELECT form_type, tb_code FROM company_forms WHERE company_id = ?", (payload.company_id,))
        forms_map = {row['form_type']: row['tb_code'] for row in forms_cursor.fetchall()}
    logging.info(f"Found company: {company_name}")

    # Authenticate with Google Drive
    logging.info("Authenticating with Google Drive.")
    try:
        drive_service = gd.get_drive_service()
    except FileNotFoundError as e:
        logging.error(f"Credential file not found: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logging.error(f"Google Drive authentication failed: {e}")
        raise HTTPException(status_code=500, detail=f"Google Drive authentication failed: {e}")
    logging.info("Google Drive authentication successful.")
    logging.info(f"Using company folder with id: {company_folder_id}")

    # --- Get Data Workflow (Same as before) ---
    logging.info("Starting Get Data Workflow.")

    # Find top-level folders within the company folder
    bank_folder_name = f"Bank_{payload.year}"
    logging.info(f"Searching for bank folder: {bank_folder_name}")
    bank_folders_query = f"'{company_folder_id}' in parents and name contains '{bank_folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
    bank_folders = gd.find_files(drive_service, bank_folders_query)
    
    logging.info("Searching for PP30 folder.")
    pp30_folders_query = f"'{company_folder_id}' in parents and name contains 'ภพ30' and mimeType = 'application/vnd.google-apps.folder'"
    pp30_folders = gd.find_files(drive_service, pp30_folders_query)

    logging.info("Searching for PND folder.")
    pnd_folders_query = f"'{company_folder_id}' in parents and name contains 'ภงด' and mimeType = 'application/vnd.google-apps.folder'"
    pnd_folders = gd.find_files(drive_service, pnd_folders_query)

    # Find files within the bank folder
    bank_files_map = {}
    if bank_folders:
        bank_folder_id = bank_folders[0]['id']
        logging.info(f"Bank folder found with id: {bank_folder_id}, searching for files.")
        for bank in banks:
            bank_name = bank['bank_name']
            search_term = bank_name
            bank_files_query = f"'{bank_folder_id}' in parents and name contains '{payload.year}{payload.month}' and name contains '{search_term}' and mimeType = 'application/pdf'"
            found_files = gd.find_files(drive_service, bank_files_query)
            bank_files_map[bank_name] = found_files
            logging.info(f"For {search_term}, found {len(found_files)} files.")

    # Find files within the PP30 folder
    pp30_files = []
    if pp30_folders:
        pp30_folder_id = pp30_folders[0]['id']
        logging.info(f"PP30 folder found with id: {pp30_folder_id}, searching for files.")
        pp30_files_query = f"'{pp30_folder_id}' in parents and name contains '{payload.year}{payload.month}' and mimeType = 'application/pdf'"
        pp30_files = gd.find_files(drive_service, pp30_files_query)
    logging.info(f"Found {len(pp30_files)} PP30 files.")

    # Find files within the PND subfolders
    sso_files, pnd1_files, pnd3_files, pnd53_files = [], [], [], []
    if pnd_folders:
        pnd_folder_id = pnd_folders[0]['id']
        logging.info(f"PND folder found with id: {pnd_folder_id}")
        def find_in_pnd_subfolder(subfolder_name):
            logging.info(f"Searching for PND subfolder: {subfolder_name}")
            subfolder_query = f"'{pnd_folder_id}' in parents and name contains '{subfolder_name}' and mimeType = 'application/vnd.google-apps.folder'"
            subfolders = gd.find_files(drive_service, subfolder_query)
            if subfolders:
                subfolder_id = subfolders[0]['id']
                logging.info(f"{subfolder_name} subfolder found with id: {subfolder_id}, searching for files.")
                file_query = f"'{subfolder_id}' in parents and name contains '{payload.year}{payload.month}' and mimeType = 'application/pdf'"
                files = gd.find_files(drive_service, file_query)
                logging.info(f"Found {len(files)} files in {subfolder_name} subfolder.")
                return files
            logging.warning(f"{subfolder_name} subfolder not found.")
            return []
        sso_files = find_in_pnd_subfolder("SSO")
        pnd1_files = find_in_pnd_subfolder("PND1")
        pnd3_files = find_in_pnd_subfolder("PND3")
        pnd53_files = find_in_pnd_subfolder("PND53")

    # Find VAT files
    vat_folder_name = f"VAT_{payload.year}"
    logging.info(f"Searching for VAT folder: {vat_folder_name}")
    vat_folders_query = f"'{company_folder_id}' in parents and name contains '{vat_folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
    vat_folders = gd.find_files(drive_service, vat_folders_query)
    vat_files = []
    if vat_folders:
        vat_folder_id = vat_folders[0]['id']
        logging.info(f"VAT folder found with id: {vat_folder_id}, searching for files.")
        year_short = str(payload.year)[-2:]
        vat_file_search_term = f"VAT{payload.month}_{year_short}"
        vat_files_query = f"'{vat_folder_id}' in parents and name contains '{vat_file_search_term}' and (mimeType = 'application/vnd.ms-excel' or mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')"
        vat_files = gd.find_files(drive_service, vat_files_query)
    logging.info(f"Found {len(vat_files)} VAT files: {[f['name'] for f in vat_files]}")

    # --- Process Data Workflow ---
    logging.info("Starting Process Data Workflow.")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Workflow Result"
    sheet.append(['Name', 'TB Code', 'File Found', 'PDF Actual Amount', 'TB Code Amount', 'Excel Actual Column', 'Result 1', 'Result 2']) # Add new header
    logging.info("Excel workbook created with new header.")

    # Find and read the TB file
    tb_files_query = f"'{company_folder_id}' in parents and name contains 'tb' and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
    tb_files = gd.find_files(drive_service, tb_files_query)
    tb_data = {}

    correct_tb_file = None
    if tb_files:
        suffix = f"{payload.year}{payload.month}.xlsx"
        for f in tb_files:
            if f['name'].endswith(suffix):
                correct_tb_file = f
                break
    
    if correct_tb_file:
        tb_file_id = correct_tb_file['id']
        logging.info(f"TB file found: {correct_tb_file['name']} (id: {tb_file_id})")
        request = drive_service.files().get_media(fileId=tb_file_id)
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)
        tb_wb = openpyxl.load_workbook(fh)
        tb_sheet = tb_wb.active
        for row in tb_sheet.iter_rows(min_row=2):
            tb_code = row[0].value
            if tb_code:
                # Safely convert cell values to float, defaulting to 0.0 on error
                try:
                    val_col7 = float(row[6].value) if len(row) > 6 and row[6].value is not None else 0.0
                except (ValueError, TypeError):
                    val_col7 = 0.0
                
                try:
                    val_col8 = float(row[7].value) if len(row) > 7 and row[7].value is not None else 0.0
                except (ValueError, TypeError):
                    val_col8 = 0.0

                amount = 0.0
                if val_col7 > 0:
                    amount = val_col7
                elif val_col8:
                    amount = -val_col8
                tb_data[str(tb_code)] = amount
    else:
        logging.warning(f"TB file for {payload.year}{payload.month} not found.")

    # Process VAT files for "Excel Actual Column"
    vat_amounts = {}
    form_mapping = {
        "ภงด.1": "PND1",
        "ภงด.3": "PND3",
        "ภงด.53": "PND53",
        "ภพ.30": "PP30",
    }
    if vat_files:
        for vat_file in vat_files:
            vat_file_id = vat_file['id']
            logging.info(f"Processing VAT file: {vat_file['name']} (id: {vat_file_id})")
            request = drive_service.files().get_media(fileId=vat_file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh.seek(0)
            df = pd.read_excel(fh, header=None)

            # Get SSO amount from cell D14
            sso_amount = "N/A"
            if df.shape[0] >= 14 and df.shape[1] >= 4:
                sso_cell_value = df.iloc[13, 3] # D14 is at index (13, 3)
                if pd.notna(sso_cell_value) and sso_cell_value != '-':
                    try:
                        sso_amount = float(sso_cell_value)
                    except (ValueError, TypeError):
                        sso_amount = "Invalid Value"
            vat_amounts["SSO"] = sso_amount
            
            for index, row in df.iterrows():
                if len(row) > 3:
                    cell_value = str(row.iloc[1])
                    for key, form_name in form_mapping.items():
                        if key in cell_value:
                            amount_cell = row.iloc[3]
                            amount = 0
                            if amount_cell != "-" and amount_cell is not None:
                                try:
                                    amount = float(amount_cell)
                                except (ValueError, TypeError):
                                    amount = "Invalid Value"
                            vat_amounts[form_name] = amount
                            # Do not break here, continue to find other forms in the file

    # Add bank data
    for bank in banks:
        bank_name = bank['bank_name']
        tb_code = bank['tb_code']
        found_files = bank_files_map.get(bank_name, [])
        file_names = ", ".join([f['name'] for f in found_files])
        
        amount = "N/A"
        if found_files:
            total_amount = 0.0
            file_count = 0
            for f in found_files:
                request = drive_service.files().get_media(fileId=f['id'])
                fh = BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                fh.seek(0)
                amount_str = get_amount_from_gemini(fh.getvalue(), PROMPTS["BANK"])
                amount_float = clean_and_convert_to_float(amount_str)
                if amount_float is not None:
                    total_amount += amount_float
                    file_count += 1
            # If multiple files are found, we sum them up.
            amount = total_amount if file_count > 0 else "N/A"
            
        tb_amount = tb_data.get(str(tb_code), "Not Found")
        sheet.append([bank_name, tb_code, file_names, amount, tb_amount, "N/A"])

        # Add formulas for result columns
        row_num = sheet.max_row
        tb_code_str = str(tb_code)
        if tb_code_str.startswith('1'):
            sheet[f'G{row_num}'] = f'=IF(D{row_num}=E{row_num}, "Correct", "Incorrect")'
            sheet[f'H{row_num}'] = "N/A"
        elif tb_code_str.startswith('2'):
            sheet[f'G{row_num}'] = f'=IF(D{row_num}=-E{row_num}, "Correct", "Incorrect")'
            sheet[f'H{row_num}'] = "N/A" # No Excel Actual Column for banks
        else:
            sheet[f'G{row_num}'] = "N/A"
            sheet[f'H{row_num}'] = "N/A"

    # Add form data
    form_data_map = {
        "PND1": (pnd1_files, forms_map.get("PND1", ""), PROMPTS.get("PND1")),
        "PND3": (pnd3_files, forms_map.get("PND3", ""), PROMPTS.get("PND3")),
        "PND53": (pnd53_files, forms_map.get("PND53", ""), PROMPTS.get("PND53")),
        "PP30": (pp30_files, forms_map.get("PP30", ""), PROMPTS.get("PP30")),
        "SSO": (sso_files, forms_map.get("SSO", ""), PROMPTS.get("SSO")),
    }

    for form_name, (files, tb_code, prompt) in form_data_map.items():
        file_names = ", ".join([f['name'] for f in files])
        amount = "N/A"
        if files and prompt:
            total_amount = 0.0
            file_count = 0
            for f in files:
                request = drive_service.files().get_media(fileId=f['id'])
                fh = BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    status, done = downloader.next_chunk()
                fh.seek(0)
                amount_str = get_amount_from_gemini(fh.getvalue(), prompt)
                amount_float = clean_and_convert_to_float(amount_str)
                if amount_float is not None:
                    total_amount += amount_float
                    file_count += 1
            amount = total_amount if file_count > 0 else "N/A"

        tb_amount = tb_data.get(str(tb_code), "Not Found")
        excel_actual_amount = vat_amounts.get(form_name, "N/A")
        sheet.append([form_name, tb_code, file_names, amount, tb_amount, excel_actual_amount])

        # Add formulas for result columns
        row_num = sheet.max_row
        tb_code_str = str(tb_code)
        if tb_code_str.startswith('1'):
            sheet[f'G{row_num}'] = f'=IF(D{row_num}=E{row_num}, "Correct", "Incorrect")'
            sheet[f'H{row_num}'] = "N/A"
        elif tb_code_str.startswith('2'):
            sheet[f'G{row_num}'] = f'=IF(D{row_num}=-E{row_num}, "Correct", "Incorrect")'
            sheet[f'H{row_num}'] = f'=IF(-E{row_num}=F{row_num}, "Correct", "Incorrect")'
        else:
            sheet[f'G{row_num}'] = "N/A"
            sheet[f'H{row_num}'] = "N/A"
    
    logging.info("Data added to the sheet.")

    # Save and return workbook
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    filename = f"{company_name}_{payload.year}{payload.month}_workflow.xlsx"
    logging.info(f"Returning Excel file: {filename}")
    
    # Properly encode filename for HTTP headers
    encoded_filename = quote(filename)
    
    return StreamingResponse(
        virtual_workbook,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )

def _calculate_monthly_totals(gl_rows: list, tb_code: str) -> dict:
    """
    Calculates the monthly totals for a given TB code from GL data.
    """
    monthly_totals = {m: 0 for m in range(1, 13)}
    in_correct_block = False
    header_skipped = False

    # This logic assumes a specific structure for the GL file.
    # It looks for a row containing the tb_code to start processing.
    # It then looks for a header row with "ลำดับที่" to start summing data.
    # It stops when it finds the next account block.

    for row in gl_rows:
        cell_a_str = str(row[0]) if row and row[0] else ""

        if not in_correct_block and tb_code in cell_a_str:
            in_correct_block = True
            continue

        if in_correct_block:
            # Stop if we hit the header of the *next* block
            if "ลำดับที่" in cell_a_str and header_skipped:
                break 
            
            # Skip the header row of the *current* block
            if "ลำดับที่" in cell_a_str and not header_skipped:
                header_skipped = True
                continue

            if header_skipped:
                # Assuming date is in the second column (index 1)
                if row and len(row) > 1 and hasattr(row[1], 'month'):
                    date = row[1]
                    month = date.month
                    
                    debit = 0.0
                    credit = 0.0

                    # Assuming Debit is in col F (idx 5) and Credit is in col G (idx 6)
                    if len(row) > 5 and row[5] is not None:
                        try:
                            debit = float(row[5])
                        except (ValueError, TypeError):
                            pass

                    if len(row) > 6 and row[6] is not None:
                        try:
                            credit = float(row[6])
                        except (ValueError, TypeError):
                            pass
                    
                    # For PP30, revenue is typically credit, and credit notes are debit.
                    # The difference (credit - debit) should give the net change.
                    amount = credit - debit
                    monthly_totals[month] += amount

    return monthly_totals


def _calculate_totals_from_sheet_data(rows: list) -> dict:
    """
    Calculates the monthly totals from a sheet of transactions.
    Assumes the sheet contains only data rows.
    """
    monthly_totals = {m: 0 for m in range(1, 13)}

    for row in rows:
        if not any(row):
            continue

        if row and len(row) > 2 and row[2]:
            try:
                date = datetime.strptime(str(row[2]), "%d/%m/%Y")
                month = date.month
            except (ValueError, TypeError):
                continue # Skip row if date is not in the expected format
            
            debit = 0.0
            credit = 0.0

            if len(row) > 6 and row[6] is not None:
                try:
                    debit = float(row[6])
                except (ValueError, TypeError):
                    pass

            if len(row) > 7 and row[7] is not None:
                try:
                    credit = float(row[7])
                except (ValueError, TypeError):
                    pass
            
            amount = credit - debit
            monthly_totals[month] += amount

    return monthly_totals


# ---------- Reconcile endpoints ----------
@app.post("/reconcile/start")
def start_reconcile(payload: ReconcileStart):
    logging.info(f"1. Reconcile started for company_id: {payload.company_id}, Year: {payload.year}, Parts: {payload.parts}")
    
    # Get company name
    with get_conn() as conn:
        logging.info("2. Fetching company information from database.")
        c = conn.execute("SELECT name, google_drive_folder_id FROM companies WHERE id = ?", (payload.company_id,)).fetchone()
        if not c:
            logging.error(f"Company with id {payload.company_id} not found.")
            raise HTTPException(status_code=404, detail="Company not found.")
        company_name = c["name"]
        company_folder_id = c["google_drive_folder_id"]
        logging.info(f"2.1. Found Company: '{company_name}' with folder ID: {company_folder_id}")
        if not company_folder_id:
            raise HTTPException(status_code=400, detail="Company does not have a Google Drive folder configured.")
        
        logging.info("2.2. Fetching company form configurations.")
        forms_cursor = conn.execute("SELECT form_type, tb_code FROM company_forms WHERE company_id = ?", (payload.company_id,))
        forms_map = {row['form_type']: row['tb_code'] for row in forms_cursor.fetchall()}
        logging.info("2.3. Form configurations loaded.")

    # Authenticate with Google Drive
    logging.info("3. Authenticating with Google Drive.")
    try:
        drive_service = gd.get_drive_service()
    except FileNotFoundError as e:
        logging.error(f"Google Drive authentication failed: Credential file not found: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        logging.error(f"Google Drive authentication failed: {e}")
        raise HTTPException(status_code=500, detail=f"Google Drive authentication failed: {e}")
    logging.info("3.1. Google Drive authentication successful.")

    # Create a new workbook for reconcile result
    logging.info("4. Creating a new Excel workbook for the reconcile result.")
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    logging.info("4.1. Workbook created.")

    # --- TB Sub-sheet ---
    if "tb_subsheet" in payload.parts:
        logging.info("5. [TB Sub-sheet] Starting TB sub-sheet creation.")
        logging.info("5.1. Looking for TB file in Google Drive.")
        tb_files_query = f"'{company_folder_id}' in parents and name contains 'tb' and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        tb_files = gd.find_files(drive_service, tb_files_query)
        
        correct_tb_file = None
        if tb_files:
            suffix = f"{payload.year}{payload.month}.xlsx"
            for f in tb_files:
                if f['name'].endswith(suffix):
                    correct_tb_file = f
                    break

        if not correct_tb_file:
            logging.error("[TB Sub-sheet] TB file not found.")
            raise HTTPException(status_code=404, detail="TB file not found for TB Sub-sheet.")

        tb_file_id = correct_tb_file['id']
        logging.info(f"5.2. Found TB file with id: {tb_file_id}. Downloading and loading.")
        request = drive_service.files().get_media(fileId=tb_file_id)
        fh = BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
        fh.seek(0)
        tb_wb = openpyxl.load_workbook(fh)
        tb_sheet = tb_wb.active
        logging.info("5.3. TB file loaded. Creating 'TB' sheet in the result workbook.")

        sheet = wb.create_sheet(title="TB")
        logging.info("5.4. Copying data from source TB to result 'TB' sheet.")
        for row_index, row in enumerate(tb_sheet.iter_rows(max_row=tb_sheet.max_row - 1), start=1):
            for col_index, cell in enumerate(row[:8], start=1):
                sheet.cell(row=row_index + 5, column=col_index, value=cell.value)
        logging.info("5.5. Data copied. Adding formulas and static values.")
        
        # Add formulas and static values
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']:
            sheet[f'{col}5'] = f'=SUBTOTAL(9,{col}8:{col}{sheet.max_row})'
        sheet.merge_cells('I6:J6')
        sheet['I6'] = 'ปรับปรุง'
        sheet['I7'] = 'เดบิท'
        sheet['J7'] = 'เครดิต'
        sheet.merge_cells('K6:K7')
        sheet['K6'] = 'Net'
        sheet['K4'] = 'กำไร'
        for row in range(8, sheet.max_row + 1):
            sheet[f'K{row}'] = f'=G{row}+I{row}-H{row}-J{row}'
        sheet['L3'] = '=+L5+L4'
        sheet['L4'] = '=+M5-L5'
        sheet['M3'] = '=+M5+M4'
        sheet.merge_cells('L6:M6')
        sheet['L6'] = 'งบกำไรขาดทุน'
        sheet['L7'] = 'เดบิท'
        sheet['M7'] = 'เครดิต'
        sheet['N3'] = '=+N5+N4'
        sheet['O2'] = '=+N3-O3'
        sheet['O3'] = '=+O5+O4'
        sheet['O4'] = '=+L4'
        for row in range(8, sheet.max_row + 1):
            a_val = sheet[f'A{row}'].value
            if a_val and str(a_val).startswith(('4', '5')):
                sheet[f'L{row}'] = f'=IF(K{row}>0,K{row},0)'
                sheet[f'M{row}'] = f'=IF(K{row}<0,-K{row},0)'
        for row in range(8, sheet.max_row + 1):
            a_val = sheet[f'A{row}'].value
            if a_val and str(a_val).startswith(('1', '2', '3')):
                sheet[f'N{row}'] = f'=IF(K{row}>0,K{row},0)'
                sheet[f'O{row}'] = f'=IF(K{row}<0,-K{row},0)'
        sheet['M2'] = '=+L3-M3'
        sheet.merge_cells('N6:O6')
        sheet['N6'] = 'งบแสดงฐานะการเงิน'
        sheet['N7'] = 'เดบิท'
        sheet['O7'] = 'เครดิต'
        sheet['A1'] = 'ชื่อบริษัท'
        sheet['B1'] = company_name
        sheet['A2'] = 'งบทดลอง ณ วันที่'
        sheet['A3'] = 'เลขที่บัญชีจาก'
        sheet['A4'] = 'วันที่จาก'
        sheet['A5'] = 'เลือกแผนก'
        sheet['B3'] = 'xxxxxx - xxxxxx'
        sheet['B5'] = '* รวมบัญชียอดเป็น 0 N'
        sheet['B4'] = '_ ถึง _'
        sheet['B2'] = f'{THAI_MONTHS[int(payload.month) - 1]} {payload.year}'
        logging.info("5.6. [TB Sub-sheet] Finished TB sub-sheet creation.")

    # --- GL-related Parts ---
    if any(part in payload.parts for part in ["gl_subsheet", "tb_code_subsheets", "pp30_subsheet"]):
        logging.info("6. [GL Parts] GL-related parts requested. Looking for GL file.")
        gl_files_query = f"'{company_folder_id}' in parents and name contains 'gl' and mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"
        gl_files = gd.find_files(drive_service, gl_files_query)
        
        correct_gl_file = None
        if gl_files:
            suffix = f"{payload.year}{payload.month}.xlsx"
            for f in gl_files:
                if f['name'].endswith(suffix):
                    correct_gl_file = f
                    break

        if correct_gl_file:
            gl_file_id = correct_gl_file['id']
            logging.info(f"6.1. Found GL file with id: {gl_file_id}. Downloading and loading.")
            request = drive_service.files().get_media(fileId=gl_file_id)
            fh = BytesIO()
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
            fh.seek(0)
            gl_wb = openpyxl.load_workbook(fh, data_only=True, read_only=True)
            gl_sheet = gl_wb.active
            logging.info("6.2. GL file loaded. Reading all rows into memory.")
            gl_rows = list(gl_sheet.iter_rows(values_only=True))
            logging.info("6.2.1. All GL rows loaded into memory.")

            if "gl_subsheet" in payload.parts:
                logging.info("7. [GL Sub-sheet] Starting GL sub-sheet creation.")
                gl_ws = wb.create_sheet(title="GL")
                logging.info("7.1. Copying all data from source GL to result 'GL' sheet.")
                for row_values in gl_rows:
                    gl_ws.append(row_values)
                logging.info("7.2. [GL Sub-sheet] Finished GL sub-sheet creation.")

            # --- tb_code_subsheets Part ---
            if "tb_code_subsheets" in payload.parts:
                logging.info("8. [TB Code Sub-sheets] Starting TB code sub-sheet creation.")
                i = 0
                account_data_block = []
                current_account_number = None
                logging.info("8.1. Iterating through GL rows to extract data for each account number.")

                while i < len(gl_rows):
                    row = gl_rows[i]
                    row_str = str(row[0]) if row and row[0] else ""

                    if "ลำดับที่" in row_str and i > 0:
                        if current_account_number and account_data_block:
                            logging.info(f"8.2. Found end of block for account '{current_account_number}'. Writing data to sheet.")
                            if current_account_number not in wb.sheetnames:
                                ws = wb.create_sheet(title=current_account_number)
                            else:
                                ws = wb[current_account_number]
                            for data_row in account_data_block:
                                ws.append(data_row)
                            ws.append([])
                        
                        account_data_block = []
                        current_account_number = None
                        
                        prev_row = gl_rows[i-1]
                        account_info = str(prev_row[0]) if prev_row and prev_row[0] else ""
                        if account_info.startswith(('1', '2')):
                            current_account_number = account_info.split()[0]
                            logging.info(f"8.3. Found start of new account block: '{current_account_number}'.")

                    if current_account_number:
                        account_data_block.append(row)
                    
                    i += 1

                if current_account_number and account_data_block:
                    logging.info(f"8.4. Writing final account block for account '{current_account_number}'.")
                    if current_account_number not in wb.sheetnames:
                        ws = wb.create_sheet(title=current_account_number)
                    else:
                        ws = wb[current_account_number]
                    for data_row in account_data_block:
                        ws.append(data_row)
                logging.info("8.5. [TB Code Sub-sheets] Finished TB code sub-sheet creation.")

            # --- pp30_subsheet Part (from GL data) ---
            if "pp30_subsheet" in payload.parts:
                logging.info("9. [PP30 Sub-sheet] Starting PP30 sub-sheet creation from GL data.")
                revenue_tb_code = forms_map.get("Revenue")
                revenue2_tb_code = forms_map.get("Revenue2")
                credit_note_tb_code = forms_map.get("Credit Note")
                logging.info(f"9.2. Found Revenue TB code: {revenue_tb_code}, Revenue2 TB code: {revenue2_tb_code}, Credit Note TB code: {credit_note_tb_code}")
                
                # --- Extract Data Blocks ---
                logging.info("9.3. Extracting data blocks from GL file.")
                
                # Extract Revenue 1
                if revenue_tb_code:
                    for i, row in enumerate(gl_rows):
                        row_str = str(row[0]) if row and row[0] else ""
                        if revenue_tb_code in row_str:
                            logging.info("9.4. Found 'รายได้' data block. Extracting data.")
                            if "รายได้" not in wb.sheetnames:
                                ws = wb.create_sheet(title="รายได้")
                            else:
                                ws = wb["รายได้"]
                            
                            block_i = i + 2
                            while block_i < len(gl_rows):
                                block_row = gl_rows[block_i]
                                if not (block_row and block_row[0] is not None and str(block_row[0]).strip()):
                                    break
                                ws.append(block_row)
                                block_i += 1

                # Extract Revenue 2
                if revenue2_tb_code:
                    for i, row in enumerate(gl_rows):
                        row_str = str(row[0]) if row and row[0] else ""
                        if revenue2_tb_code in row_str:
                            logging.info("9.4.1. Found 'รายได้ 2' data block. Extracting data.")
                            if "รายได้ 2" not in wb.sheetnames:
                                ws = wb.create_sheet(title="รายได้ 2")
                            else:
                                ws = wb["รายได้ 2"]

                            block_i = i + 2
                            while block_i < len(gl_rows):
                                block_row = gl_rows[block_i]
                                if not (block_row and block_row[0] is not None and str(block_row[0]).strip()):
                                    break
                                ws.append(block_row)
                                block_i += 1

                # Extract Credit Note
                if credit_note_tb_code:
                    for i, row in enumerate(gl_rows):
                        row_str = str(row[0]) if row and row[0] else ""
                        if credit_note_tb_code in row_str:
                            logging.info("9.5. Found 'ลดหนี้' data block. Extracting data.")
                            if "ลดหนี้" not in wb.sheetnames:
                                ws = wb.create_sheet(title="ลดหนี้")
                            else:
                                ws = wb["ลดหนี้"]
                            
                            block_i = i + 2
                            while block_i < len(gl_rows):
                                block_row = gl_rows[block_i]
                                if not (block_row and block_row[0] is not None and str(block_row[0]).strip()):
                                    break
                                ws.append(block_row)
                                block_i += 1
                
                logging.info("9.8. Calculating monthly totals from 'รายได้', 'รายได้ 2', and 'ลดหนี้' sheets.")
                revenue_totals = {}
                if "รายได้" in wb.sheetnames:
                    revenue_rows = list(wb["รายได้"].iter_rows(values_only=True))
                    revenue_totals = _calculate_totals_from_sheet_data(revenue_rows)

                revenue2_totals = {}
                if "รายได้ 2" in wb.sheetnames:
                    revenue2_rows = list(wb["รายได้ 2"].iter_rows(values_only=True))
                    revenue2_totals = _calculate_totals_from_sheet_data(revenue2_rows)

                credit_note_totals = {}
                if "ลดหนี้" in wb.sheetnames:
                    credit_note_rows = list(wb["ลดหนี้"].iter_rows(values_only=True))
                    credit_note_totals = _calculate_totals_from_sheet_data(credit_note_rows)

                logging.info("9.9. Monthly totals calculated. Populating 'PP30' sheet.")

                if "PP30" not in wb.sheetnames:
                    pp30_ws = wb.create_sheet(title="PP30")
                    pp30_ws['C4'] = "เดือน"
                    pp30_ws['D4'] = "PP30"
                    pp30_ws['E4'] = "รายได้"
                    pp30_ws['F4'] = "รายได้ 2"
                    pp30_ws['G4'] = "ลดหนี้"
                    pp30_ws['H4'] = "Diff"
                    thai_months = ["มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน", "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"]
                    for i, month in enumerate(thai_months):
                        pp30_ws[f'C{i+5}'] = month
                
                pp30_ws = wb["PP30"]
                for month_num in range(1, 13):
                    pp30_ws[f'E{month_num+4}'] = revenue_totals.get(month_num, 0) if revenue_totals.get(month_num, 0) != 0 else "-"
                    pp30_ws[f'F{month_num+4}'] = revenue2_totals.get(month_num, 0) if revenue2_totals.get(month_num, 0) != 0 else "-"
                    pp30_ws[f'G{month_num+4}'] = credit_note_totals.get(month_num, 0) if credit_note_totals.get(month_num, 0) != 0 else "-"
                logging.info("9.10. [PP30 Sub-sheet] Finished populating with GL data.")

        else:
            logging.warning("6. [GL Parts] GL file not found. Skipping GL-dependent parts.")

    # --- PP30 Sub-sheet (from PDF data) ---
    if "pp30_subsheet" in payload.parts:
        logging.info("10. [PP30 Sub-sheet] Starting to populate PP30 data from monthly PDF files.")
        
        pp30_ws = wb["PP30"] if "PP30" in wb.sheetnames else None

        if not pp30_ws:
            logging.info("10.1. 'PP30' sheet not found, creating it now.")
            pp30_ws = wb.create_sheet(title="PP30")
            # ... (headers)
        
        logging.info("10.2. Looking for 'ภพ30' folder in Google Drive.")
        pp30_folders_query = f"'{company_folder_id}' in parents and name contains 'ภพ30' and mimeType = 'application/vnd.google-apps.folder'"
        pp30_folders = gd.find_files(drive_service, pp30_folders_query)
        
        if pp30_folders:
            pp30_folder_id = pp30_folders[0]['id']
            logging.info(f"10.3. Found 'ภพ30' folder with ID: {pp30_folder_id}. Iterating through months 1-12.")
            for i, month in enumerate(range(1, 13)):
                month_str = f"{month:02d}"
                logging.info(f"10.4. Searching for PP30 PDF for month {month_str}.")
                file_query = f"'{pp30_folder_id}' in parents and name contains '{payload.year}{month_str}' and mimeType = 'application/pdf'"
                pp30_files = gd.find_files(drive_service, file_query)

                if pp30_files:
                    file_id = pp30_files[0]['id']
                    logging.info(f"10.5. Found PDF '{pp30_files[0]['name']}'. Downloading and processing with Gemini.")
                    request = drive_service.files().get_media(fileId=file_id)
                    fh = BytesIO()
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()
                    fh.seek(0)
                    
                    prompt = "จากเอกสารนี้ ให้ดึงตัวเลขของหัวข้อ 'ยอดขายที่ต้องเสียภาษี' ออกมา ตอบกลับเฉพาะตัวเลขเท่านั้น ห้ามมีตัวหนังสือเด็ดขาด"
                    amount_str = get_amount_from_gemini(fh.getvalue(), prompt)
                    amount = clean_and_convert_to_float(amount_str)
                    logging.info(f"10.6. Gemini extracted: '{amount_str}', Converted to: {amount}")
                    
                    if amount is None:
                        logging.warning(f"Could not convert '{amount_str}' to a number for PP30 month {month_str}.")
                        amount = amount_str
                    
                    if pp30_ws:
                        pp30_ws[f'D{i+5}'] = amount if amount is not None else "-"
                else:
                    logging.info(f"10.5. No PP30 PDF found for month {month_str}.")
                    if pp30_ws:
                        pp30_ws[f'D{i+5}'] = "-"
        else:
            logging.warning("10.3. 'ภพ30' folder not found. Skipping PDF data extraction for PP30.")
            for i in range(12):
                if pp30_ws:
                    pp30_ws[f'D{i+5}'] = "-"
        logging.info("10.7. [PP30 Sub-sheet] Finished populating with PDF data.")
    
    # Save and return workbook
    logging.info("11. Saving final workbook to a virtual file.")
    virtual_workbook = BytesIO()
    wb.save(virtual_workbook)
    virtual_workbook.seek(0)
    filename = f"{company_name}_{payload.year}{payload.month}_reconcile.xlsx"
    logging.info(f"11.1. Returning Excel file: {filename}")

    encoded_filename = quote(filename)
    return StreamingResponse(
        virtual_workbook,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )
