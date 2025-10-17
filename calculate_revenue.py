import openpyxl
import sys
from datetime import datetime

def calculate_sheet_monthly_totals(rows: list) -> dict:
    """
    Calculates the monthly totals from a sheet of transactions.
    Assumes the sheet contains a header row with "ลำดับที่" and then data rows.
    """
    monthly_totals = {m: 0 for m in range(1, 13)}

    for row in rows:
        # Skip empty rows
        if not any(row):
            continue
        
        if row and len(row) > 1:
            print("---")
            date_ = datetime.strptime(row[2], "%d/%m/%Y")
            month = date_.month
            
            debit = 0.0
            credit = 0.0

            # Assuming Debit is in col F (idx 5) and Credit is in col G (idx 6)
            if len(row) > 6 and row[6] is not None:
                try:
                    debit = float(row[6])
                except (ValueError, TypeError):
                    pass

            if len(row) > 7 and row[7] is not None:
                try:
                    credit = float(row[7])
                except (ValueError, TypeError):
                    pass
            
            # The difference (credit - debit) should give the net change.
            amount = credit - debit
            monthly_totals[month] += amount

    return monthly_totals

def main():
    if len(sys.argv) != 3:
        print("Usage: python calculate_revenue.py <path_to_excel_file> <sheet_name>")
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2]

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        if sheet_name not in workbook.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
            print(f"Available sheets: {workbook.sheetnames}")
            sys.exit(1)
            
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        
        print(f"Calculating monthly totals from sheet '{sheet_name}'...")
        monthly_totals = calculate_sheet_monthly_totals(rows)
        
        print("\n--- Monthly Totals ---")
        for month, total in monthly_totals.items():
            print(f"Month {month:02d}: {total:,.2f}")
        print("----------------------\n")

    except FileNotFoundError:
        print(f"Error: File not found at '{file_path}'")
        sys.exit(1)
    except Exception as e:
        print(f"An error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()